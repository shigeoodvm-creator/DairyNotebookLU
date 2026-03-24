"""
CalfScan 7葉方式への更新スクリプト
- 個体記録票: 列ヘッダーをR2-R6/L2-L6 → 葉①〜⑦ に変更
- スコア判定ガイド: 7葉定義セクション追加（出典付き）
- プローブ位置ガイド: イラスト再生成（7葉対応）
"""

import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
import copy
import sys

sys.stdout.reconfigure(encoding='utf-8')

# =============================
# 共通スタイルヘルパー
# =============================

def solid_fill(hex_color):
    return PatternFill(fill_type='solid', fgColor=hex_color)

def make_font(bold=False, color='FF000000', size=9, name='MS Gothic'):
    return Font(bold=bold, color=color, size=size, name=name)

def make_align(h='center', v='center', wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border():
    s = Side(style='thin', color='FF999999')
    return Border(left=s, right=s, top=s, bottom=s)

# =============================
# 葉定義（CalfScan準拠）
# =============================

LOBE_DEFS = [
    # (葉番号, 短縮名, 日本語名, 英語名, 走査ICS, 側, fill_row5, fill_row6)
    (1, '葉①\nR前-前', '右前葉・前部',
     'Cranial part of right cranial lobe', '右 ICS 1〜2',
     'R', 'FF21618C', 'FF2E86C1'),
    (2, '葉②\nR前-後', '右前葉・後部',
     'Caudal part of right cranial lobe', '右 ICS 3〜4',
     'R', 'FF21618C', 'FF2E86C1'),
    (3, '葉③\nR中', '右中葉',
     'Right middle lobe', '右 ICS 5',
     'R', 'FF21618C', 'FF2E86C1'),
    (4, '葉④\nR後', '右後葉',
     'Right caudal lobe', '右 ICS 6〜10',
     'R', 'FF21618C', 'FF2E86C1'),
    (5, '葉⑤\nL前-前', '左前葉・前部',
     'Cranial part of left cranial lobe', '左 ICS 2〜3',
     'L', 'FF1B5E20', 'FF27AE60'),
    (6, '葉⑥\nL前-後', '左前葉・後部',
     'Caudal part of left cranial lobe', '左 ICS 4〜5',
     'L', 'FF1B5E20', 'FF27AE60'),
    (7, '葉⑦\nL後', '左後葉',
     'Left caudal lobe', '左 ICS 6〜10',
     'L', 'FF1B5E20', 'FF27AE60'),
]

# Excelの列マッピング（F=葉①, G=葉②, H=葉③, I=葉④, J=スペーサー, K=葉⑤, L=葉⑥, M=葉⑦）
LOBE_COL = {1: 'F', 2: 'G', 3: 'H', 4: 'I', 5: 'K', 6: 'L', 7: 'M'}
SPACER_COLS = ['J', 'N', 'O']

# =============================
# メイン処理
# =============================

WB_PATH = r'C:\Users\user\OneDrive\デスクトップ\肺エコープロジェクト\肺エコー検診_記録様式.xlsx'
wb = openpyxl.load_workbook(WB_PATH)

# ===============================
# 1. 個体記録票 シートの更新
# ===============================
ws = wb['個体記録票']

# --- 既存のマージを解除してから再設定 ---
# F5:J5, K5:O5 の結合を解除
for merge_range in ['F5:J5', 'K5:O5']:
    try:
        ws.unmerge_cells(merge_range)
    except Exception:
        pass

# 右肺ヘッダー: F5:I5 を結合
ws.merge_cells('F5:I5')
ws['F5'].value = '右肺　各肺葉所見（CalfScan 葉①〜④）\n（○=正常Aライン  B=Bライン増多  ●=実質化  E=胸水  A=膿瘍）'
ws['F5'].fill = solid_fill('FF21618C')
ws['F5'].font = make_font(bold=True, color='FFFFFFFF', size=8)
ws['F5'].alignment = make_align('center', 'center', True)

# J5 スペーサー
ws['J5'].value = ''
ws['J5'].fill = solid_fill('FFD6EAF8')

# 左肺ヘッダー: K5:M5 を結合
ws.merge_cells('K5:M5')
ws['K5'].value = '左肺　各肺葉所見（CalfScan 葉⑤〜⑦）\n（○=正常Aライン  B=Bライン増多  ●=実質化  E=胸水  A=膿瘍）'
ws['K5'].fill = solid_fill('FF1B5E20')
ws['K5'].font = make_font(bold=True, color='FFFFFFFF', size=8)
ws['K5'].alignment = make_align('center', 'center', True)

# N5, O5 スペーサー
for col in ['N', 'O']:
    ws[f'{col}5'].value = ''
    ws[f'{col}5'].fill = solid_fill('FFD5F5E3')

# --- Row 6: 各肺葉ラベル ---
ws.row_dimensions[6].height = 30  # 2行テキスト用に高さ拡張

for lobe_num, short_name, jp_name, en_name, ics, side, fill5, fill6 in LOBE_DEFS:
    col = LOBE_COL[lobe_num]
    cell = ws[f'{col}6']
    cell.value = short_name
    cell.fill = solid_fill(fill6)
    cell.font = make_font(bold=True, color='FFFFFFFF', size=7)
    cell.alignment = make_align('center', 'center', True)
    cell.border = thin_border()

    # Excelコメント（全情報）
    comment_text = (
        f'【葉{lobe_num}】{jp_name}\n'
        f'{en_name}\n'
        f'走査範囲: {ics}\n\n'
        f'出典: CalfScan App, UW-Madison\n'
        f'Ollivett & Buczinski, Vet Clin NA Food Anim 2016'
    )
    cell.comment = Comment(comment_text, 'CalfScan定義')

# スペーサー列のrow6
for col in SPACER_COLS:
    ws[f'{col}6'].value = ''
    fill_color = 'FFD6EAF8' if col == 'J' else 'FFD5F5E3'
    ws[f'{col}6'].fill = solid_fill(fill_color)

# --- Row 5 高さ調整 ---
ws.row_dimensions[5].height = 32

# --- Row 4 スコア早見の "3: 1葉性" 等の注記を更新 ---
# スコア早見行はrow4（結合セルあり）。値はA4:C4の結合セルに入っている想定
# ※元テキストを確認しながら葉の定義注記を追加
score_note_col = None
for col_idx in range(1, ws.max_column + 1):
    cell = ws.cell(row=4, column=col_idx)
    val = cell.value
    if val and '5: 3葉以上' in str(val):
        # この後ろに注記追加
        cell.value = str(val) + '  ※CalfScan7葉方式（前葉前部・後部は各1葉）'
        break

# =============================
# 2. スコア判定ガイドシートへ CalfScan 7葉定義を追加
# =============================
ws_guide = wb['スコア判定ガイド']
max_r = ws_guide.max_row

# 空行 + セクション開始
start_row = max_r + 2

# セクションタイトル
ws_guide.merge_cells(f'A{start_row}:H{start_row}')
title_cell = ws_guide[f'A{start_row}']
title_cell.value = '【CalfScan 7葉方式  ─  肺葉定義・走査範囲（本記録様式の準拠方式）】'
title_cell.fill = solid_fill('FF1A5276')
title_cell.font = make_font(bold=True, color='FFFFFFFF', size=10)
title_cell.alignment = make_align('left', 'center', False)
ws_guide.row_dimensions[start_row].height = 20

# 副タイトル
sub_row = start_row + 1
ws_guide.merge_cells(f'A{sub_row}:H{sub_row}')
sub_cell = ws_guide[f'A{sub_row}']
sub_cell.value = (
    '出典: Ollivett TL, Buczinski S. On-farm use of ultrasonography for bovine respiratory disease. '
    'Vet Clin North Am Food Anim Pract. 2016;32(1):19-35. doi:10.1016/j.cvfa.2015.09.001  '
    '／  CalfScan App, University of Wisconsin-Madison School of Veterinary Medicine '
    '（https://www.vetmed.wisc.edu/fapm/svm-dairy-apps/calfscan-calf-lung-ultrasound-scorer/）'
)
sub_cell.fill = solid_fill('FFD6EAF8')
sub_cell.font = make_font(bold=False, color='FF1A237E', size=8)
sub_cell.alignment = make_align('left', 'center', True)
ws_guide.row_dimensions[sub_row].height = 28

# テーブルヘッダー
hdr_row = sub_row + 1
headers = ['記録票\n列', '葉\n番号', '日本語名', '英語名（原著・CalfScan）', '走査肋間\n(ICS)', '側', 'スコア計算時の\n「1葉」扱い', '備考']
col_widths = [6, 6, 16, 36, 12, 5, 16, 20]
hdr_fills = ['FF2E86C1', 'FF2E86C1', 'FF2E86C1', 'FF2E86C1',
             'FF2E86C1', 'FF2E86C1', 'FF2E86C1', 'FF2E86C1']

for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
    cell = ws_guide.cell(row=hdr_row, column=ci)
    cell.value = h
    cell.fill = solid_fill('FF2E4057')
    cell.font = make_font(bold=True, color='FFFFFFFF', size=8)
    cell.alignment = make_align('center', 'center', True)
    cell.border = thin_border()
    ws_guide.column_dimensions[get_column_letter(ci)].width = w

ws_guide.row_dimensions[hdr_row].height = 28

# テーブル行
table_data = [
    # col, 葉番, 日本語, 英語, ICS, 側, スコア, 備考
    ('F', '葉①', '右前葉・前部\n(cranial部)', 'Cranial part of right cranial lobe', '右 ICS 1〜2', '右', '1葉として計上', ''),
    ('G', '葉②', '右前葉・後部\n(caudal部)', 'Caudal part of right cranial lobe', '右 ICS 3〜4', '右', '1葉として計上', '前部・後部は\n独立した別葉'),
    ('H', '葉③', '右中葉', 'Right middle lobe', '右 ICS 5', '右', '1葉として計上', ''),
    ('I', '葉④', '右後葉', 'Right caudal lobe', '右 ICS 6〜10', '右', '1葉として計上', ''),
    ('K', '葉⑤', '左前葉・前部\n(cranial部)', 'Cranial part of left cranial lobe', '左 ICS 2〜3', '左', '1葉として計上', ''),
    ('L', '葉⑥', '左前葉・後部\n(caudal部)', 'Caudal part of left cranial lobe', '左 ICS 4〜5', '左', '1葉として計上', '前部・後部は\n独立した別葉'),
    ('M', '葉⑦', '左後葉', 'Left caudal lobe', '左 ICS 6〜10', '左', '1葉として計上', '右副葉は\n評価対象外'),
]

row_colors_r = ['FFEAF4FB', 'FFD6EAF8', 'FFAED6F1', 'FF7FB3D3']
row_colors_l = ['FFE9F7EF', 'FFD5F5E3', 'FFA9DFBF', 'FF7DCEA0']

for i, row_data in enumerate(table_data):
    dr = hdr_row + 1 + i
    side = row_data[5]
    base_color = row_colors_r[i % 4] if side == '右' else row_colors_l[i % 4]
    for ci, val in enumerate(row_data, start=1):
        cell = ws_guide.cell(row=dr, column=ci)
        cell.value = val
        cell.fill = solid_fill(base_color)
        cell.font = make_font(bold=(ci == 2), color='FF1A1A2E', size=8)
        cell.alignment = make_align('center' if ci in (1, 2, 5, 6) else 'left', 'center', True)
        cell.border = thin_border()
    ws_guide.row_dimensions[dr].height = 30

# スコア計算の説明ボックス
note_start = hdr_row + 1 + len(table_data) + 1
ws_guide.merge_cells(f'A{note_start}:H{note_start}')
note_cell = ws_guide[f'A{note_start}']
note_cell.value = (
    '【スコア3〜5における「葉数」の数え方（CalfScan方式）】\n'
    'スコア3 = 上記7葉のうち 1葉が完全実質化（hepatization）\n'
    'スコア4 = 2葉が完全実質化\n'
    'スコア5 = 3葉以上が完全実質化\n'
    '※ 右前葉の前部（葉①）と後部（葉②）、左前葉の前部（葉⑤）と後部（葉⑥）は、それぞれ独立した「1葉」としてカウントする。\n'
    '   5葉システム（前葉前部+後部をまとめて1葉とする方式）と混在しないよう注意。'
)
note_cell.fill = solid_fill('FFFFFDE7')
note_cell.font = make_font(bold=False, color='FF4A1000', size=8)
note_cell.alignment = make_align('left', 'top', True)
ws_guide.row_dimensions[note_start].height = 80

# =============================
# 保存
# =============================
wb.save(WB_PATH)
print('✓ 個体記録票・スコア判定ガイド 更新完了')

# =============================
# 3. プローブ位置ガイドイラスト再生成（7葉対応）
# =============================
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import numpy as np

plt.rcParams['font.family'] = 'MS Gothic'
plt.rcParams['axes.unicode_minus'] = False

LOBE_COLORS = {
    1: '#E74C3C', 2: '#E67E22', 3: '#3498DB', 4: '#1ABC9C',
    5: '#9B59B6', 6: '#E91E63', 7: '#00897B'
}

LOBE_ICS_R = {1: (1, 2), 2: (3, 4), 3: (5, 5), 4: (6, 10)}
LOBE_ICS_L = {5: (2, 3), 6: (4, 5), 7: (6, 10)}

LOBE_LABEL_R = {1: '葉①\n右前-前', 2: '葉②\n右前-後', 3: '葉③\n右中', 4: '葉④\n右後'}
LOBE_LABEL_L = {5: '葉⑤\n左前-前', 6: '葉⑥\n左前-後', 7: '葉⑦\n左後'}


def draw_body(ax, flip=False):
    sx = -1 if flip else 1
    body = mpatches.Ellipse((0, 0), width=5.2, height=2.8,
                             facecolor='#F5E6C8', edgecolor='#8B6914', linewidth=2.0, zorder=2)
    ax.add_patch(body)
    nx = [sx*2.0, sx*2.6, sx*3.1, sx*3.3, sx*2.9, sx*2.2, sx*2.0]
    ny = [0.8, 1.2, 1.4, 1.0, 0.5, 0.4, 0.8]
    ax.fill(nx, ny, color='#F5E6C8', zorder=3)
    ax.plot(nx, ny, color='#8B6914', linewidth=2.0, zorder=4)
    head = mpatches.Ellipse((sx*3.4, 1.05), width=1.1, height=0.85,
                             facecolor='#F5E6C8', edgecolor='#8B6914', linewidth=2.0, zorder=4)
    ax.add_patch(head)
    nose = mpatches.Ellipse((sx*3.95, 0.85), width=0.3, height=0.25,
                             facecolor='#E8C9A0', edgecolor='#8B6914', linewidth=1.5, zorder=5)
    ax.add_patch(nose)
    ax.plot([sx*3.35], [1.25], 'o', color='#4A3000', markersize=5, zorder=6)
    ear_x = [sx*3.1, sx*3.0, sx*3.3, sx*3.3, sx*3.1]
    ear_y = [1.4, 1.75, 1.65, 1.3, 1.4]
    ax.fill(ear_x, ear_y, color='#F5E6C8', zorder=3)
    ax.plot(ear_x, ear_y, color='#8B6914', linewidth=1.5, zorder=4)
    fl_x = sx * 1.6
    for (x0, y_top, y_bot, w) in [(fl_x, -0.3, -1.0, 0.22), (fl_x, -1.0, -2.3, 0.18)]:
        ax.fill([x0-w, x0+w, x0+w, x0-w], [y_top, y_top, y_bot, y_bot],
                color='#F5E6C8', zorder=2)
        ax.plot([x0-w, x0-w, x0+w, x0+w, x0-w], [y_top, y_bot, y_bot, y_top, y_top],
                color='#8B6914', linewidth=1.5, zorder=3)
    elbow = mpatches.Circle((fl_x, -1.0), radius=0.12, facecolor='#FFD700',
                              edgecolor='#B8860B', linewidth=1.5, zorder=5)
    ax.add_patch(elbow)
    rl_x = sx * (-1.6)
    for (x0, y_top, y_bot, w) in [(rl_x, -0.3, -1.0, 0.22), (rl_x, -1.0, -2.3, 0.18)]:
        ax.fill([x0-w, x0+w, x0+w, x0-w], [y_top, y_top, y_bot, y_bot],
                color='#F5E6C8', zorder=2)
        ax.plot([x0-w, x0-w, x0+w, x0+w, x0-w], [y_top, y_bot, y_bot, y_top, y_top],
                color='#8B6914', linewidth=1.5, zorder=3)
    return fl_x


def draw_lobes(ax, flip=False):
    sx = -1 if flip else 1
    rib_start_x = sx * 1.55
    rib_spacing = sx * (-0.35)
    lung_top = 0.88
    lung_bot = -0.85
    n_ribs = 11

    rib_xs = [rib_start_x + rib_spacing * i for i in range(n_ribs)]

    # 各肺葉ゾーンを描画
    lobe_ics = LOBE_ICS_R if not flip else LOBE_ICS_L
    lobe_labels = LOBE_LABEL_R if not flip else LOBE_LABEL_L

    for lobe_num, (ics_start, ics_end) in lobe_ics.items():
        # ICSのindexは肋間番号に対応: ICS1 = rib_xs[0]〜rib_xs[1] の間
        xi_start = ics_start - 1  # ICS1 の左側 rib index
        xi_end = ics_end          # ICS_end の右側 rib index
        if xi_start >= len(rib_xs) or xi_end >= len(rib_xs):
            xi_end = min(xi_end, len(rib_xs) - 1)
        x1 = rib_xs[xi_start]
        x2 = rib_xs[xi_end]
        xmin, xmax = min(x1, x2), max(x1, x2)
        color = LOBE_COLORS[lobe_num]
        # 塗りつぶし
        rect = mpatches.FancyBboxPatch(
            (xmin + 0.01, lung_bot - 0.04), xmax - xmin - 0.02,
            lung_top - lung_bot + 0.08,
            boxstyle='round,pad=0.02',
            facecolor=color, alpha=0.22,
            edgecolor=color, linewidth=1.5, zorder=3)
        ax.add_patch(rect)
        # ラベル（上部）
        mx = (xmin + xmax) / 2
        ax.text(mx, lung_top + 0.18, lobe_labels[lobe_num],
                ha='center', va='bottom', fontsize=6.5, color=color,
                fontweight='bold', zorder=7,
                bbox=dict(boxstyle='round,pad=0.15', facecolor='white',
                          edgecolor=color, alpha=0.9, linewidth=1))
        # プローブマーカー
        probe_y = 0.05
        pw, ph = 0.07, 0.16
        probe = mpatches.FancyBboxPatch(
            (mx - pw, probe_y - ph/2), pw*2, ph,
            boxstyle='round,pad=0.01',
            facecolor=color, edgecolor='white', linewidth=1.5, zorder=7, alpha=0.92)
        ax.add_patch(probe)
        ax.annotate('', xy=(mx, probe_y - ph/2 - 0.14),
                    xytext=(mx, probe_y + ph/2 + 0.1),
                    arrowprops=dict(arrowstyle='->', color=color, lw=1.8), zorder=8)

    # 肋骨ライン
    for i, rx in enumerate(rib_xs[:10]):
        ax.plot([rx, rx], [lung_bot - 0.12, lung_top + 0.04],
                color='#C0A070', linewidth=1.5, zorder=4, alpha=0.6)
        if i < 9:
            ics_n = i + 1
            ax.text((rx + rib_xs[i+1])/2, lung_bot - 0.18,
                    f'{ics_n}', ha='center', va='top', fontsize=6,
                    color='#888888', zorder=5)


fig, axes = plt.subplots(1, 2, figsize=(17, 7.5))
fig.patch.set_facecolor('#F8F9FA')

configs = [
    (False, '右側面 (Right side)', '右肺'),
    (True,  '左側面 (Left side)',  '左肺'),
]

for ax, (flip, title, side_label) in zip(axes, configs):
    ax.set_facecolor('#EEF4FF')
    ax.set_xlim(-5.2, 5.2)
    ax.set_ylim(-3.2, 3.2)
    ax.set_aspect('equal')
    ax.axis('off')
    ax.set_title(title, fontsize=11, fontweight='bold', color='#1A1A2E', pad=6)

    fl_x = draw_body(ax, flip=flip)
    draw_lobes(ax, flip=flip)

    sx = -1 if flip else 1
    ax.annotate('肘関節\n(前葉前部の\n後方から開始)',
                xy=(fl_x, -1.0), xytext=(fl_x + sx*(-0.7), -1.7),
                fontsize=6.5, ha='center', color='#7D6608', fontweight='bold',
                arrowprops=dict(arrowstyle='->', color='#B8860B', lw=1.5),
                bbox=dict(boxstyle='round,pad=0.2', facecolor='#FFF9E6',
                          edgecolor='#B8860B', alpha=0.95), zorder=9)

    ax.text(0, -2.95,
            '数字は肋間番号（ICS）。各色ゾーン＝各肺葉。プローブ▼は背側→腹側方向にスライド。',
            ha='center', fontsize=7, color='#555555', style='italic', zorder=10)

# 凡例（全7葉）
lobe_names_r = {1: '葉① 右前葉・前部', 2: '葉② 右前葉・後部', 3: '葉③ 右中葉', 4: '葉④ 右後葉'}
lobe_names_l = {5: '葉⑤ 左前葉・前部', 6: '葉⑥ 左前葉・後部', 7: '葉⑦ 左後葉'}

legend_patches = []
for ln, name in {**lobe_names_r, **lobe_names_l}.items():
    legend_patches.append(
        mpatches.Patch(facecolor=LOBE_COLORS[ln], edgecolor='white',
                       label=name, alpha=0.85))

fig.legend(handles=legend_patches, loc='lower center', ncol=4,
           fontsize=8, framealpha=0.95,
           title='CalfScan 7葉定義（Ollivett & Buczinski 2016 / UW-Madison CalfScan App準拠）',
           title_fontsize=8.5,
           bbox_to_anchor=(0.5, 0.01))

fig.suptitle('子牛 肺エコー検診  プローブ当て位置ガイド  ─  CalfScan 7葉方式',
             fontsize=13, fontweight='bold', color='#1A1A2E', y=0.99)

plt.tight_layout(rect=[0, 0.10, 1, 0.97])
img_path = r'C:\Users\user\OneDrive\デスクトップ\肺エコープロジェクト\肺エコー_プローブ位置ガイド.png'
plt.savefig(img_path, dpi=150, bbox_inches='tight', facecolor=fig.get_facecolor())
plt.close()
print('✓ イラスト再生成完了')

# =============================
# 4. プローブ位置ガイドシートの画像を差し替え
# =============================
from openpyxl.drawing.image import Image as XLImage

wb2 = openpyxl.load_workbook(WB_PATH)

# 既存の「プローブ位置ガイド」シートを削除して再作成
if 'プローブ位置ガイド' in [s.title for s in wb2.worksheets]:
    del wb2['プローブ位置ガイド']

ws_pg = wb2.create_sheet('プローブ位置ガイド', 1)

# タイトル
ws_pg['A1'] = '子牛 肺エコー検診  プローブ当て位置ガイド  ─  CalfScan 7葉方式'
ws_pg['A1'].font = Font(bold=True, size=11, name='MS Gothic')
ws_pg['A1'].fill = PatternFill(fill_type='solid', fgColor='FF1A5276')
ws_pg['A1'].font = Font(bold=True, size=11, color='FFFFFFFF', name='MS Gothic')
ws_pg.row_dimensions[1].height = 22
ws_pg.merge_cells('A1:J1')

# 出典
ws_pg['A2'] = ('出典: Ollivett TL, Buczinski S. Vet Clin North Am Food Anim Pract. 2016;32(1):19-35. '
               '/ CalfScan App, UW-Madison School of Veterinary Medicine')
ws_pg['A2'].font = Font(size=8, italic=True, name='MS Gothic', color='FF1A237E')
ws_pg['A2'].fill = PatternFill(fill_type='solid', fgColor='FFD6EAF8')
ws_pg.merge_cells('A2:J2')
ws_pg.row_dimensions[2].height = 14

# 画像挿入
img = XLImage(img_path)
img.width = 960
img.height = 430
ws_pg.add_image(img, 'A3')

# 個体記録票の右側のサムネも差し替え
ws_rec = wb2['個体記録票']
# 既存の画像を削除（openpyxlでは直接削除できないため、新規ファイル保存時に反映される）
# 代わりにW1を使わずV1に新規挿入
img2 = XLImage(img_path)
img2.width = 520
img2.height = 232
ws_rec.add_image(img2, 'V1')

wb2.save(WB_PATH)
print('✓ プローブ位置ガイドシート・個体記録票サムネ 更新完了')
print('\n=== 全処理完了 ===')
